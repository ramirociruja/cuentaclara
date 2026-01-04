# app/services/onboarding_import_validate.py
from __future__ import annotations
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple, Optional
from datetime import datetime, timedelta, timezone
import io

from openpyxl import load_workbook

REQUIRED_SHEETS = ["Customers", "Loans", "Payments"]

CUSTOMERS_REQUIRED = ["customer_ref", "first_name", "last_name"]
LOANS_REQUIRED = ["loan_ref", "customer_ref", "amount", "total_due", "installments_count", "installment_amount", "frequency"]
PAYMENTS_REQUIRED = ["payment_ref", "loan_ref", "amount"]

ALLOWED_FREQUENCY = {"weekly", "monthly"}
ALLOWED_PAYMENT_TYPE = {"cash", "transfer", "other"}
ALLOWED_LOAN_STATUS = {"active", "paid", "defaulted"}

@dataclass
class RowIssue:
    sheet: str
    row: int
    field: str
    code: str
    message: str

def _norm_header(s: Any) -> str:
    return str(s or "").strip()

def _as_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def _as_float(v: Any) -> Optional[float]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def _as_int(v: Any) -> Optional[int]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None

def _as_date(v: Any) -> Optional[datetime]:
    # openpyxl puede traer datetime ya parseado
    if isinstance(v, datetime):
        if v.tzinfo is None:
            return v.replace(tzinfo=timezone.utc)
        return v.astimezone(timezone.utc)
    if v is None or str(v).strip() == "":
        return None
    # si te pasan string "YYYY-MM-DD" o similar, lo resolvemos simple:
    try:
        # intenta ISO completo
        dt = datetime.fromisoformat(str(v).strip())
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    except Exception:
        return None

def _read_sheet_rows(wb, sheet_name: str) -> Tuple[List[Dict[str, Any]], List[RowIssue]]:
    issues: List[RowIssue] = []
    if sheet_name not in wb.sheetnames:
        issues.append(RowIssue(sheet_name, 0, "", "MISSING_SHEET", f"Falta la hoja {sheet_name}"))
        return [], issues

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        issues.append(RowIssue(sheet_name, 0, "", "EMPTY_SHEET", "Hoja vacía"))
        return [], issues

    headers = [_norm_header(h) for h in header_row]
    header_map: Dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h:
            header_map[h] = idx

    data: List[Dict[str, Any]] = []
    excel_row_index = 1  # header is 1
    for r in rows_iter:
        excel_row_index += 1
        if r is None:
            continue

        # descartar fila completamente vacía
        if all((c is None or str(c).strip() == "") for c in r):
            continue

        row_obj: Dict[str, Any] = {}
        for h, idx in header_map.items():
            row_obj[h] = r[idx] if idx < len(r) else None
        row_obj["__rownum__"] = excel_row_index
        data.append(row_obj)

    return data, issues

def validate_onboarding_xlsx(file_bytes: bytes) -> Dict[str, Any]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)

    errors: List[RowIssue] = []
    warnings: List[RowIssue] = []

    # check sheets
    for s in REQUIRED_SHEETS:
        if s not in wb.sheetnames:
            errors.append(RowIssue(s, 0, "", "MISSING_SHEET", f"Falta la hoja obligatoria: {s}"))

    if errors:
        return _build_result([], [], [], errors, warnings)

    customers, iss = _read_sheet_rows(wb, "Customers")
    errors.extend(iss)
    loans, iss = _read_sheet_rows(wb, "Loans")
    errors.extend(iss)
    payments, iss = _read_sheet_rows(wb, "Payments")
    errors.extend(iss)

    # required columns presence (by checking first row keys)
    def _ensure_required(sheet: str, rows: List[Dict[str, Any]], required: List[str]):
        if not rows:
            errors.append(RowIssue(sheet, 0, "", "EMPTY_SHEET", "No hay filas de datos"))
            return
        cols = set([k for k in rows[0].keys() if not k.startswith("__")])
        for c in required:
            if c not in cols:
                errors.append(RowIssue(sheet, 1, c, "MISSING_COLUMN", f"Falta columna obligatoria: {c}"))

    _ensure_required("Customers", customers, CUSTOMERS_REQUIRED)
    _ensure_required("Loans", loans, LOANS_REQUIRED)
    _ensure_required("Payments", payments, PAYMENTS_REQUIRED)

    if errors:
        return _build_result(customers, loans, payments, errors, warnings)

    # build lookups
    customer_refs = set()
    customer_ref_to_row = {}

    for c in customers:
        rn = int(c["__rownum__"])
        cref = _as_str(c.get("customer_ref"))
        if not cref:
            errors.append(RowIssue("Customers", rn, "customer_ref", "REQUIRED", "customer_ref es obligatorio"))
            continue
        if cref in customer_refs:
            errors.append(RowIssue("Customers", rn, "customer_ref", "DUPLICATE_REF", f"customer_ref duplicado: {cref}"))
            continue
        customer_refs.add(cref)
        customer_ref_to_row[cref] = rn

        fn = _as_str(c.get("first_name"))
        ln = _as_str(c.get("last_name"))
        if not fn:
            errors.append(RowIssue("Customers", rn, "first_name", "REQUIRED", "first_name es obligatorio"))
        if not ln:
            errors.append(RowIssue("Customers", rn, "last_name", "REQUIRED", "last_name es obligatorio"))

        dni = _as_str(c.get("dni"))
        if not dni:
            warnings.append(RowIssue("Customers", rn, "dni", "MISSING", "DNI vacío; dedupe por DNI no será posible"))

    loan_refs = set()
    for l in loans:
        rn = int(l["__rownum__"])
        lref = _as_str(l.get("loan_ref"))
        if not lref:
            errors.append(RowIssue("Loans", rn, "loan_ref", "REQUIRED", "loan_ref es obligatorio"))
            continue
        if lref in loan_refs:
            errors.append(RowIssue("Loans", rn, "loan_ref", "DUPLICATE_REF", f"loan_ref duplicado: {lref}"))
            continue
        loan_refs.add(lref)

        cref = _as_str(l.get("customer_ref"))
        if not cref:
            errors.append(RowIssue("Loans", rn, "customer_ref", "REQUIRED", "customer_ref es obligatorio"))
        elif cref not in customer_refs:
            errors.append(RowIssue("Loans", rn, "customer_ref", "NOT_FOUND", f"customer_ref no existe en Customers: {cref}"))

        freq = _as_str(l.get("frequency"))
        if not freq or freq not in ALLOWED_FREQUENCY:
            errors.append(RowIssue("Loans", rn, "frequency", "INVALID", "frequency debe ser weekly o monthly"))

        amount = _as_float(l.get("amount"))
        total_due = _as_float(l.get("total_due"))
        icount = _as_int(l.get("installments_count"))
        iamount = _as_float(l.get("installment_amount"))

        if amount is None or amount <= 0:
            errors.append(RowIssue("Loans", rn, "amount", "INVALID", "amount debe ser numérico y > 0"))
        if total_due is None or total_due <= 0:
            errors.append(RowIssue("Loans", rn, "total_due", "INVALID", "total_due debe ser numérico y > 0"))
        if icount is None or icount <= 0:
            errors.append(RowIssue("Loans", rn, "installments_count", "INVALID", "installments_count debe ser entero > 0"))
        if iamount is None or iamount <= 0:
            errors.append(RowIssue("Loans", rn, "installment_amount", "INVALID", "installment_amount debe ser numérico y > 0"))

        st = _as_str(l.get("status"))
        if st and st not in ALLOWED_LOAN_STATUS:
            errors.append(RowIssue("Loans", rn, "status", "INVALID", "status debe ser active/paid/defaulted"))

        cd = _as_int(l.get("collection_day"))
        if cd is not None and not (1 <= cd <= 7):
            errors.append(RowIssue("Loans", rn, "collection_day", "INVALID", "collection_day debe ser 1..7"))

        sd = _as_date(l.get("start_date"))
        if l.get("start_date") and sd is None:
            errors.append(RowIssue("Loans", rn, "start_date", "INVALID", "start_date inválida (ISO o fecha Excel)"))

    payment_refs = set()
    for p in payments:
        rn = int(p["__rownum__"])
        pref = _as_str(p.get("payment_ref"))
        if not pref:
            errors.append(RowIssue("Payments", rn, "payment_ref", "REQUIRED", "payment_ref es obligatorio"))
            continue
        if pref in payment_refs:
            errors.append(RowIssue("Payments", rn, "payment_ref", "DUPLICATE_REF", f"payment_ref duplicado: {pref}"))
            continue
        payment_refs.add(pref)

        lref = _as_str(p.get("loan_ref"))
        if not lref:
            errors.append(RowIssue("Payments", rn, "loan_ref", "REQUIRED", "loan_ref es obligatorio"))
        elif lref not in loan_refs:
            errors.append(RowIssue("Payments", rn, "loan_ref", "NOT_FOUND", f"loan_ref no existe en Loans: {lref}"))

        amt = _as_float(p.get("amount"))
        if amt is None or amt <= 0:
            errors.append(RowIssue("Payments", rn, "amount", "INVALID", "amount debe ser numérico y > 0"))

        pd = _as_date(p.get("payment_date"))
        if p.get("payment_date") and pd is None:
            errors.append(RowIssue("Payments", rn, "payment_date", "INVALID", "payment_date inválida (ISO o fecha Excel)"))

        pt = _as_str(p.get("payment_type"))
        if pt and pt not in ALLOWED_PAYMENT_TYPE:
            errors.append(RowIssue("Payments", rn, "payment_type", "INVALID", "payment_type debe ser cash/transfer/other"))

    return _build_result(customers, loans, payments, errors, warnings)

def _build_result(customers, loans, payments, errors: List[RowIssue], warnings: List[RowIssue]) -> Dict[str, Any]:
    def issue_to_dict(i: RowIssue) -> Dict[str, Any]:
        return {"sheet": i.sheet, "row": i.row, "field": i.field, "code": i.code, "message": i.message}

    payload = {
        "customers": _normalize_customers(customers),
        "loans": _normalize_loans(loans),
        "payments": _normalize_payments(payments),
    }

    errors_list = [issue_to_dict(e) for e in errors]
    warnings_list = [issue_to_dict(w) for w in warnings]

    summary = _build_productive_summary(payload, errors_list, warnings_list)

    return {
        "payload": payload,
        "summary": summary,
        "errors": errors_list,
        "warnings": warnings_list,
    }

def _build_productive_summary(payload: Dict[str, Any], errors: list[dict], warnings: list[dict]) -> Dict[str, Any]:
    customers = payload.get("customers", [])
    loans = payload.get("loans", [])
    payments = payload.get("payments", [])

    # -------------------------
    # QUALITY
    # -------------------------
    def _count_by_sheet(items: list[dict], sheet: str) -> int:
        return sum(1 for x in items if x.get("sheet") == sheet)

    quality = {
        "blocking_errors": len(errors) > 0,
        "errors_total": len(errors),
        "warnings_total": len(warnings),
        "by_sheet": {
            "Customers": {
                "rows": len(customers),
                "errors": _count_by_sheet(errors, "Customers"),
                "warnings": _count_by_sheet(warnings, "Customers"),
            },
            "Loans": {
                "rows": len(loans),
                "errors": _count_by_sheet(errors, "Loans"),
                "warnings": _count_by_sheet(warnings, "Loans"),
            },
            "Payments": {
                "rows": len(payments),
                "errors": _count_by_sheet(errors, "Payments"),
                "warnings": _count_by_sheet(warnings, "Payments"),
            },
        },
    }

    # -------------------------
    # COVERAGE
    # -------------------------
    def _has(v) -> bool:
        return v is not None and str(v).strip() != ""

    coverage = {
        "customers": {
            "with_dni": sum(1 for c in customers if _has(c.get("dni"))),
            "with_phone": sum(1 for c in customers if _has(c.get("phone"))),
            "with_email": sum(1 for c in customers if _has(c.get("email"))),
            "with_address": sum(1 for c in customers if _has(c.get("address"))),
            "with_province": sum(1 for c in customers if _has(c.get("province"))),
        },
        "loans": {
            "with_employee_email": sum(1 for l in loans if _has(l.get("employee_email"))),
            "missing_employee_email": sum(1 for l in loans if not _has(l.get("employee_email"))),
            "with_start_date": sum(1 for l in loans if _has(l.get("start_date"))),
            "missing_start_date": sum(1 for l in loans if not _has(l.get("start_date"))),
        },
        "payments": {
            "with_collector_email": sum(1 for p in payments if _has(p.get("collector_email"))),
            "missing_collector_email": sum(1 for p in payments if not _has(p.get("collector_email"))),
            "with_payment_type": sum(1 for p in payments if _has(p.get("payment_type"))),
            "missing_payment_type": sum(1 for p in payments if not _has(p.get("payment_type"))),
            "with_payment_date": sum(1 for p in payments if _has(p.get("payment_date"))),
            "missing_payment_date": sum(1 for p in payments if not _has(p.get("payment_date"))),
        },
    }

    # -------------------------
    # CONSISTENCY (cross refs, duplicates, mismatches)
    # -------------------------
    customer_refs = [c.get("customer_ref") for c in customers if _has(c.get("customer_ref"))]
    loan_refs = [l.get("loan_ref") for l in loans if _has(l.get("loan_ref"))]
    payment_refs = [p.get("payment_ref") for p in payments if _has(p.get("payment_ref"))]

    def _count_dupes(refs: list[str]) -> int:
        seen = set()
        dupes = 0
        for r in refs:
            if r in seen:
                dupes += 1
            else:
                seen.add(r)
        return dupes

    # Dangling refs: lo detectamos en payload también (sin depender de errors)
    customer_ref_set = set(customer_refs)
    loan_ref_set = set(loan_refs)

    loans_missing_customer = sum(
        1 for l in loans
        if _has(l.get("customer_ref")) and l.get("customer_ref") not in customer_ref_set
    )
    payments_missing_loan = sum(
        1 for p in payments
        if _has(p.get("loan_ref")) and p.get("loan_ref") not in loan_ref_set
    )

    # Mismatch: installments_count * installment_amount vs total_due (tolerancia)
    tolerance = 1.0  # pesos / unidad monetaria
    mismatch_count = 0
    for l in loans:
        ic = l.get("installments_count")
        ia = l.get("installment_amount")
        td = l.get("total_due")
        if isinstance(ic, int) and isinstance(ia, (int, float)) and isinstance(td, (int, float)):
            expected = ic * float(ia)
            if abs(expected - float(td)) > tolerance:
                mismatch_count += 1

    consistency = {
        "duplicate_refs": {
            "customer_ref": _count_dupes(customer_refs),
            "loan_ref": _count_dupes(loan_refs),
            "payment_ref": _count_dupes(payment_refs),
        },
        "dangling_refs": {
            "loans_missing_customer": loans_missing_customer,
            "payments_missing_loan": payments_missing_loan,
        },
        "loan_amount_mismatches": {
            "count": mismatch_count,
            "tolerance": tolerance,
        },
    }

    # -------------------------
    # IMPACT (volumen y montos)
    # -------------------------
    installments_to_generate_total = sum(
        int(l.get("installments_count") or 0) for l in loans
        if isinstance(l.get("installments_count"), int)
    )

    loans_total_due_sum = sum(
        float(l.get("total_due") or 0) for l in loans
        if isinstance(l.get("total_due"), (int, float))
    )

    payments_total_amount = sum(
        float(p.get("amount") or 0) for p in payments
        if isinstance(p.get("amount"), (int, float))
    )

    estimated_paid_ratio = 0.0
    if loans_total_due_sum > 0:
        estimated_paid_ratio = payments_total_amount / loans_total_due_sum

    impact = {
        "customers_total": len(customers),
        "loans_total": len(loans),
        "payments_total": len(payments),
        "installments_to_generate_total": installments_to_generate_total,
        "loans_total_due_sum": round(loans_total_due_sum, 2),
        "payments_total_amount": round(payments_total_amount, 2),
        "estimated_paid_ratio": round(estimated_paid_ratio, 4),
    }

    # -------------------------
    # RISKS (banderas rojas operativas)
    # -------------------------
    customers_without_identifiers = sum(
        1 for c in customers
        if not _has(c.get("dni")) and not _has(c.get("phone")) and not _has(c.get("email"))
    )

    risks = {
        "customers_without_identifiers": customers_without_identifiers,
        "loans_without_start_date": coverage["loans"]["missing_start_date"],
        "payments_without_payment_date": coverage["payments"]["missing_payment_date"],
        "loans_without_employee_email": coverage["loans"]["missing_employee_email"],
        "payments_without_collector_email": coverage["payments"]["missing_collector_email"],
    }

    return {
        "quality": quality,
        "coverage": coverage,
        "consistency": consistency,
        "impact": impact,
        "risks": risks,
    }


def _normalize_customers(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for r in rows:
        out.append({
            "rownum": int(r["__rownum__"]),
            "customer_ref": _as_str(r.get("customer_ref")),
            "first_name": _as_str(r.get("first_name")),
            "last_name": _as_str(r.get("last_name")),
            "dni": _as_str(r.get("dni")),
            "phone": _as_str(r.get("phone")),
            "email": _as_str(r.get("email")),
            "address": _as_str(r.get("address")),
            "province": _as_str(r.get("province")),
        })
    return out

def _normalize_loans(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for r in rows:
        sd = _as_date(r.get("start_date"))
        out.append({
            "rownum": int(r["__rownum__"]),
            "loan_ref": _as_str(r.get("loan_ref")),
            "customer_ref": _as_str(r.get("customer_ref")),
            "employee_email": _as_str(r.get("employee_email")),
            "amount": _as_float(r.get("amount")),
            "total_due": _as_float(r.get("total_due")),
            "installments_count": _as_int(r.get("installments_count")),
            "installment_amount": _as_float(r.get("installment_amount")),
            "frequency": _as_str(r.get("frequency")),
            "start_date": sd.isoformat() if sd else None,
            "status": _as_str(r.get("status")) or "active",
            "description": _as_str(r.get("description")),
            "collection_day": _as_int(r.get("collection_day")),
        })
    return out

def _normalize_payments(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for r in rows:
        pd = _as_date(r.get("payment_date"))
        out.append({
            "rownum": int(r["__rownum__"]),
            "payment_ref": _as_str(r.get("payment_ref")),
            "loan_ref": _as_str(r.get("loan_ref")),
            "amount": _as_float(r.get("amount")),
            "payment_date": pd.isoformat() if pd else None,
            "payment_type": _as_str(r.get("payment_type")),
            "description": _as_str(r.get("description")),
            "collector_email": _as_str(r.get("collector_email")),
        })
    return out
